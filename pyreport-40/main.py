import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Cargar el archivo Excel existente (asegúrate de que ya existe)
archivo_excel = ""
libro = openpyxl.load_workbook(archivo_excel)
hoja = libro.active

# Obtener la última fila con datos (asumiendo que la primera fila es la cabecera)
ultima_fila = hoja.max_row

# Definir los datos del soporte técnico (puedes modificar esto según tus necesidades)
nuevo_soporte = {
    "Prioridad": "Prioridad",
    "Descripción": "Descripción",
    "Abierto el Día": "Abierto el Día",
    "Informado por": "Informado por",
    "Asignado a": "Asignado a",
    "Fecha de Resolución": "Fecha de Resolución"
}

# Agregar los datos a la siguiente fila vacía
for columna, valor in nuevo_soporte.items():
    letra_columna = get_column_letter(hoja[f"{columna}1"].column)
    hoja[f"{letra_columna}{ultima_fila + 1}"] = valor

# Formatear la nueva fila (opcional)
for celda in hoja[f"A{ultima_fila + 1}:F{ultima_fila + 1}"]:
    for c in celda:
        c.font = Font(bold=True)

# Guardar los cambios
libro.save(archivo_excel)
print("Registro de soporte técnico agregado correctamente.")
